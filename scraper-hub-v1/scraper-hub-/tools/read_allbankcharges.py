from openpyxl import load_workbook
p = r"c:\Users\USER 2\Downloads\scraper-hub-v1 (2)\AllBankCharges.xlsx"
wb = load_workbook(p, read_only=True)
ws = wb.active
for i, row in enumerate(ws.iter_rows(values_only=True)):
    print('|'.join([str(c) if c is not None else '' for c in row]))
    if i >= 50:
        break
