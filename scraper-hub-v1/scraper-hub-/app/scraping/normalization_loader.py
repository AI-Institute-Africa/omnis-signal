import os
import re
from pathlib import Path
from openpyxl import load_workbook
from typing import List, Dict

SPREADSHEET_ENV_VAR = 'BANKING_NORMALIZATION_SPREADSHEET'
SPREADSHEET_NAMES = ('AllBankCharges.xlsx', 'AllBankCharges*.xlsx')


def _resolve_search_paths() -> List[Path]:
    current = Path.cwd().resolve()
    module_root = Path(__file__).resolve().parent
    project_root = module_root.parents[2] if len(module_root.parents) >= 3 else module_root

    paths = [current]
    if project_root not in paths:
        paths.append(project_root)
    paths.extend([parent for parent in current.parents[:3]])
    paths.extend([parent for parent in project_root.parents[:3]])
    return [p for p in paths if p.exists()]


def _find_spreadsheet(names=SPREADSHEET_NAMES) -> Path:
    env_path = os.getenv(SPREADSHEET_ENV_VAR)
    if env_path:
        candidate = Path(env_path)
        if candidate.exists():
            return candidate

    for root in _resolve_search_paths():
        for pattern in names:
            hits = list(root.rglob(pattern))
            if hits:
                return hits[0]
    return None


CATEGORY_MAP = {
    'account management': 'account_maintenance',
    'transfers': 'bank_transfer',
    'payments': 'bill_payment',
    'cards': 'card_service',
    'cash services': 'cash_withdrawal',
    'lending': 'loan',
    'deposits': 'savings_account',
    'transactional banking': 'general_banking',
    'retail banking': 'general_banking',
    'business banking': 'general_banking',
    'service fees': 'account_maintenance',
}


def _normalize_subcategory(service_category: str, service_type: str, product: str) -> str:
    if service_category:
        normalized = str(service_category).strip().lower()
        mapped = CATEGORY_MAP.get(normalized)
        if mapped:
            return mapped
        if 'saving' in normalized or 'deposit' in normalized:
            return 'savings_account'
        if 'current' in normalized or 'transaction' in normalized:
            return 'current_account'
        if 'credit' in normalized or 'card' in normalized:
            return 'card_service'
        if 'loan' in normalized or 'lending' in normalized or 'overdraft' in normalized:
            return 'loan'
        if 'transfer' in normalized or 'payment' in normalized:
            return 'bank_transfer'

    raw = str(service_type or product or '').strip().lower()
    if 'transfer' in raw or 'rtgs' in raw or 'zimswitch' in raw or 'swift' in raw or 'wire' in raw:
        return 'bank_transfer'
    if 'card' in raw or 'visa' in raw or 'mastercard' in raw or 'debit' in raw or 'credit' in raw:
        return 'card_service'
    if 'loan' in raw or 'overdraft' in raw or 'lending' in raw:
        return 'loan'
    if 'withdrawal' in raw or 'cash' in raw:
        return 'cash_withdrawal'
    if 'internet' in raw or 'online' in raw or 'mobile' in raw or 'e-banking' in raw or 'sms' in raw:
        return 'internet_banking'
    if 'account maintenance' in raw or 'monthly ledger fee' in raw or 'monthly fee' in raw:
        return 'account_maintenance'
    if 'payment' in raw or 'merchant' in raw or 'bill' in raw:
        return 'bill_payment'
    if 'saving' in raw or 'deposit' in raw:
        return 'savings_account'
    if 'current' in raw or 'transaction' in raw:
        return 'current_account'
    return 'banking'


def load_banking_normalization(path: str = None) -> List[Dict]:
    """Load banking normalization mappings from an Excel spreadsheet.

    Expected columns (case-insensitive): Product, Service Type, Service Category, ChargeType, Currency, Value, Bank
    Returns a list of rule dicts compatible with BANKING_SERVICE_NORMALIZATION.
    """
    p = Path(path) if path else _find_spreadsheet()
    if not p or not p.exists():
        return []

    wb = load_workbook(str(p), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(h).strip() if h is not None else '' for h in rows[0]]
    header_lc = [h.lower() for h in header]

    def idx(name):
        try:
            return header_lc.index(name.lower())
        except ValueError:
            return None

    i_product = idx('product')
    i_service_type = idx('service type')
    i_service_category = idx('service category')
    i_charge_type = idx('chargetype') or idx('charge type')
    i_value = idx('value')
    i_currency = idx('currency')
    i_bank = idx('bank')

    rules = []
    seen = set()
    for row in rows[1:]:
        product = row[i_product] if i_product is not None else None
        service_type = row[i_service_type] if i_service_type is not None else None
        service_category = row[i_service_category] if i_service_category is not None else None
        charge_type = row[i_charge_type] if i_charge_type is not None else None
        value = row[i_value] if i_value is not None else None
        currency = row[i_currency] if i_currency is not None else None
        bank = row[i_bank] if i_bank is not None else None

        product_text = str(product).strip() if product else ''
        service_type_text = str(service_type).strip() if service_type else ''
        if not product_text and not service_type_text:
            continue

        key = (product_text, service_type_text)
        if key in seen:
            continue
        seen.add(key)

        parts = []
        if product_text:
            parts.append(re.escape(product_text))
        if service_type_text and service_type_text.lower() not in product_text.lower():
            parts.append(re.escape(service_type_text))
        if not parts:
            continue

        pattern = r"\b(" + r"|".join(parts) + r")\b"
        subcategory = _normalize_subcategory(service_category, service_type_text, product_text)
        title = product_text or service_type_text or 'Banking Service'
        item_name = product_text or service_type_text

        rules.append({
            'pattern': pattern,
            'subcategory': subcategory,
            'title': title,
            'item_name': item_name,
            'source_bank': str(bank).strip() if bank else None,
            'example_value': value,
            'currency': currency,
            'charge_type': str(charge_type).strip() if charge_type else None,
        })

    return rules
